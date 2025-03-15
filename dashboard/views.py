# vouchers/views.py
from django.views.generic import TemplateView
from django.db.models import Count, Sum
from django.utils import timezone
from datetime import timedelta
import json
from vouchers.models import voucher

class VoucherDashboardView(TemplateView):
    template_name = 'voucher_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Obtenha os filtros de data dos parâmetros GET
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        vouchers_qs = voucher.objects.all()

        if start_date:
            vouchers_qs = vouchers_qs.filter(data__date__gte=start_date)
        if end_date:
            vouchers_qs = vouchers_qs.filter(data__date__lte=end_date)

        # Agregações para os cards do dashboard
        total_vouchers = vouchers_qs.count()
        total_value = vouchers_qs.aggregate(total=Sum('valor'))['total'] or 0
        total_gasto = vouchers_qs.aggregate(total=Sum('gasto'))['total'] or 0

        # Dados para o gráfico: quantidade de vouchers por status
        status_data = vouchers_qs.values('status').annotate(count=Count('id'))

        # Dados para contagem por embarcador
        embarcador_data = vouchers_qs.values('embarcador').annotate(count=Count('id'))

        context['total_vouchers'] = total_vouchers
        context['total_value'] = total_value
        context['total_gasto'] = total_gasto
        context['status_labels'] = json.dumps([item['status'] for item in status_data])
        context['status_counts'] = json.dumps([item['count'] for item in status_data])
        context['embarcador_labels'] = json.dumps(
            [item['embarcador'] if item['embarcador'] else "Sem Embarcador" for item in embarcador_data]
        )
        context['embarcador_counts'] = json.dumps([item['count'] for item in embarcador_data])
        context['vouchers'] = vouchers_qs.order_by('-data')  # Ordena por data decrescente
        context['start_date'] = start_date if start_date else ''
        context['end_date'] = end_date if end_date else ''
        return context
    
    # dashboard/views.py

from django.http import HttpResponse
from django.utils import timezone
from django.shortcuts import get_list_or_404
import io
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

import matplotlib
matplotlib.use('Agg')  # Uso 'Agg' para gerar imagens em memória (sem interface)
import matplotlib.pyplot as plt
import base64

from vouchers.models import voucher
from datetime import datetime
# views.py
from django.http import HttpResponse
from datetime import datetime
import io
import openpyxl

# Mantendo seu model em minúsculo (não é recomendável, mas funciona)
from vouchers.models import voucher

def download_excel_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    vouchers_qs = voucher.objects.all()
    # Filtros de data, se necessários
    # ...

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório"

    headers = ["ID", "Motorista", "Placa", "Embarcador", "Rota", "Valor", "Status", "Telefone", "Data"]
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_num, value=header)

    for row_num, v in enumerate(vouchers_qs, start=2):
        ws.cell(row=row_num, column=1, value=v.id)
        ws.cell(row=row_num, column=2, value=v.motorista)
        ws.cell(row=row_num, column=3, value=v.placa)
        ws.cell(row=row_num, column=4, value=str(v.embarcador) if v.embarcador else "")
        
        # Ajuste aqui para converter a Rota em string:
        ws.cell(row=row_num, column=5, value=str(v.rota) if v.rota else "")
        
        ws.cell(row=row_num, column=6, value=float(v.valor) if v.valor else 0.0)
        ws.cell(row=row_num, column=7, value=v.status)
        ws.cell(row=row_num, column=8, value=v.telefone)

        data_str = v.data.strftime("%d/%m/%Y %H:%M") if v.data else ""
        ws.cell(row=row_num, column=9, value=data_str)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="relatorio_vouchers.xlsx"'
    return response

  
def download_pdf_report(request):
    """Gera e faz o download de um relatório PDF com a tabela de vouchers e os gráficos."""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Filtra vouchers
    vouchers = voucher.objects.all()
    if start_date:
        vouchers = vouchers.filter(data__gte=start_date)
    if end_date:
        dt_end = datetime.strptime(end_date, "%Y-%m-%d")
        dt_end = dt_end.replace(hour=23, minute=59, second=59)
        vouchers = vouchers.filter(data__lte=dt_end)

    # --- PREPARANDO OS DADOS PARA OS GRÁFICOS ---
    # 1) Gráfico de Status (Barra)
    status_counts = {}
    for v in vouchers:
        status_counts[v.status] = status_counts.get(v.status, 0) + 1

    # 2) Gráfico de Embarcador (Pizza)
    embarcador_counts = {}
    for v in vouchers:
        embarcador_counts[v.embarcador] = embarcador_counts.get(v.embarcador, 0) + 1

    # --- CRIANDO IMAGENS MATPLOTLIB EM MEMÓRIA ---

    # Gráfico de Status (Barra)
    fig1, ax1 = plt.subplots(figsize=(4, 3))  # Tamanho do gráfico (largura=4, altura=3)
    ax1.bar(status_counts.keys(), status_counts.values(), color='skyblue')
    ax1.set_title("Distribuição de Vouchers por Status")
    ax1.set_xlabel("Status")
    ax1.set_ylabel("Quantidade")
    # Salva em memória
    buf1 = io.BytesIO()
    plt.savefig(buf1, format='png')
    buf1.seek(0)
    plt.close(fig1)  # Fecha o gráfico para liberar memória

    # Gráfico de Embarcador (Pizza)
    fig2, ax2 = plt.subplots(figsize=(4, 3))
    labels = list(embarcador_counts.keys())
    sizes = list(embarcador_counts.values())
    ax2.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=140)
    ax2.set_title("Distribuição de Vouchers por Embarcador")
    # Salva em memória
    buf2 = io.BytesIO()
    plt.savefig(buf2, format='png')
    buf2.seek(0)
    plt.close(fig2)

    # --- AGORA CRIAMOS O PDF COM REPORTLAB ---

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)

    styles = getSampleStyleSheet()
    story = []

    # Título
    story.append(Paragraph("Relatório de Vouchers", styles['Title']))
    story.append(Spacer(1, 12))
    # Intervalo de datas
    story.append(Paragraph(f"Data Início: {start_date or 'N/A'}", styles['Normal']))
    story.append(Paragraph(f"Data Fim: {end_date or 'N/A'}", styles['Normal']))
    story.append(Spacer(1, 12))

    # Inserindo os gráficos
    # 1) Status
    img_status = Image(buf1)
    img_status._restrictSize(200, 150)  # Ajuste de tamanho
    story.append(img_status)
    story.append(Spacer(1, 12))
    # 2) Embarcador
    img_embarcador = Image(buf2)
    img_embarcador._restrictSize(200, 150)
    story.append(img_embarcador)
    story.append(Spacer(1, 24))

    # Tabela com todos os vouchers
    data_table = []
    # Cabeçalho
    data_table.append(["ID", "Motorista", "Placa", "Embarcador", "Rota", "Valor", "Status", "Telefone", "Data"])
    
    for v in vouchers:
        data_str = v.data.strftime("%d/%m/%Y %H:%M") if v.data else ""
        row = [
            v.id,
            v.motorista,
            v.placa,
            v.embarcador,
            v.rota,
            f"R$ {v.valor:.2f}" if v.valor else "R$ 0.00",
            v.status,
            v.telefone,
            data_str
        ]
        data_table.append(row)

    # Cria a tabela
    t = Table(data_table, repeatRows=1)  # repeatRows=1 repete cabeçalho a cada página
    t.setStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.gray),
        ('TEXTCOLOR',(0, 0),(-1, 0), colors.white),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
    ])
    story.append(t)

    doc.build(story)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio_vouchers.pdf"'
    return response
